import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Input and output paths
input_file = r"C:\Users\foendoe.kevin\Documents\Excel\Cassy\Maturity_A\Maturity_A.xlsx"
output_dir = r"C:\Users\foendoe.kevin\Documents\Excel\Cassy\Maturity - Send to Companies"

# Read Excel sheets
df_contracts = pd.read_excel(input_file, sheet_name='CMP_Contract_Maturities')
df_members = pd.read_excel(input_file, sheet_name='CMP_Member_Input_Details')

# Extract contract columns
contract_col = [col for col in df_contracts.columns if 'CONTRACT' in col][0]
member_contract_col = [col for col in df_members.columns if 'CONTRACT' in col][0]

# Ensure output directory exists
os.makedirs(output_dir, exist_ok=True)

# Function to clean column names
def extract_column_name(col_name):
    match = re.search(r'\[(.*?)\]', col_name)
    return match.group(1) if match else col_name

# Final required column order (without 'middle')
final_columns = [
    'Contract No', 'planname', 'Mbr No', 'lastname', 'firstname', 'birthdt',
    'gender', 'email_address', 'PhoneNumber', 'Phone Type', 'national_id',
    'Address line 1', 'Type', 'Address line 2', 'Address line 3', 'Post Code',
    'City', 'State', 'Country'
]

# Mapping from original to final column names (no 'middle')
column_mapping = {
    'CONTRACT': 'Contract No',
    'PLANNAME': 'planname',
    'MBR_NO': 'Mbr No',
    'LASTNAME': 'lastname',
    'FIRSTNAME': 'firstname',
    'BIRTHDT': 'birthdt',
    'GENDER': 'gender',
    'EMAIL_ADDRESS': 'email_address',
    'DIAL_NUMBER': 'PhoneNumber',
    'NATIONAL_ID': 'national_id',
    'ADDRESS_UNIQUE': 'Address line 1'
}

# Loop through each contract
for contract in df_contracts[contract_col].unique():
    matching_rows = df_members[df_members[member_contract_col] == contract]

    if not matching_rows.empty:
        cleaned_df = matching_rows.copy()
        cleaned_df.columns = [extract_column_name(col) for col in cleaned_df.columns]

        # Rename columns
        cleaned_df = cleaned_df.rename(columns=column_mapping)

        # Add missing columns
        for col in final_columns:
            if col not in cleaned_df.columns:
                if col == 'Type' or col == 'Phone Type':
                    cleaned_df[col] = 'HOME'
                else:
                    cleaned_df[col] = ""

        # Make sure 'Type' and 'Phone Type' are set
        cleaned_df['Type'] = cleaned_df['Type'].replace('', 'HOME')
        cleaned_df['Phone Type'] = cleaned_df['Phone Type'].replace('', 'HOME')

        # Format birthdate
        if 'birthdt' in cleaned_df.columns:
            cleaned_df['birthdt'] = pd.to_datetime(cleaned_df['birthdt'], errors='coerce').dt.strftime('%d-%m-%Y')

        # Reorder columns
        cleaned_df = cleaned_df[final_columns]

        # Output file path
        filename = f"{contract} - request member info.xlsx"
        output_path = os.path.join(output_dir, filename)

        # Write to Excel
        cleaned_df.to_excel(output_path, index=False)

        # Apply styles
        wb = load_workbook(output_path)
        ws = wb.active

        header_font = Font(bold=True, size=16)
        data_font = Font(size=16)
        fill_gray = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        fill_header = PatternFill(start_color="D1F0FF", end_color="D1F0FF", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill_header

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = data_font
                if cell.row % 2 == 0:
                    cell.fill = fill_gray

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    font_size = cell.font.size or 16
                    char_width = font_size / 11
                    cell_length = len(str(cell.value)) * char_width
                    max_length = max(max_length, cell_length)
            ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

        wb.save(output_path)
